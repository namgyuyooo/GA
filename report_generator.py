"""
Excel 보고서 생성 모듈
데이터 시각화 및 포맷팅 포함
"""

import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
import config


def create_excel_report(utm_summary, channel_summary, landing_summary, 
                       search_analysis, performance_summary, insights):
    """종합 Excel 보고서 생성"""
    
    filename = f"{config.REPORT_SETTINGS['output_filename'][:-5]}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        
        # 1. 요약 시트
        create_summary_sheet(writer, performance_summary, insights)
        
        # 2. UTM 캠페인 성과
        utm_summary.to_excel(writer, sheet_name='UTM 캠페인 성과', index=False)
        
        # 3. 채널별 성과
        channel_summary.to_excel(writer, sheet_name='채널별 성과', index=False)
        
        # 4. 랜딩 페이지 성과
        landing_summary.to_excel(writer, sheet_name='랜딩 페이지 성과', index=False)
        
        # 5. 검색 키워드 성과
        search_analysis['top_keywords'].to_excel(writer, sheet_name='검색 키워드', index=False)
        
        # 6. SEO 기회 분석
        if len(search_analysis['low_ctr_opportunities']) > 0:
            search_analysis['low_ctr_opportunities'].to_excel(
                writer, sheet_name='SEO 개선 기회', index=False
            )
        
        # 7. 고성과 키워드
        if len(search_analysis['high_ctr_keywords']) > 0:
            search_analysis['high_ctr_keywords'].to_excel(
                writer, sheet_name='고성과 키워드', index=False
            )
    
    # 스타일링 적용
    apply_excel_formatting(filename)
    
    return filename


def create_summary_sheet(writer, performance_summary, insights):
    """요약 시트 생성"""
    
    # 성과 요약 DataFrame 생성
    utm_perf = performance_summary['utm_performance']
    search_perf = performance_summary['search_performance']
    
    summary_data = [
        ['📊 UTM 성과 요약', ''],
        ['총 세션', f"{utm_perf['total_sessions']:,}"],
        ['총 사용자', f"{utm_perf['total_users']:,}"],
        ['총 전환', f"{utm_perf['total_conversions']:,}"],
        ['평균 참여율', f"{utm_perf['avg_engagement_rate']:.1%}"],
        ['캠페인 수', f"{utm_perf['unique_campaigns']}개"],
        ['', ''],
        ['🔍 검색 성과 요약', ''],
        ['총 클릭', f"{search_perf['total_clicks']:,}"],
        ['총 노출', f"{search_perf['total_impressions']:,}"],
        ['평균 CTR', f"{search_perf['avg_ctr']:.2%}"],
        ['평균 순위', f"{search_perf['avg_position']:.1f}위"],
        ['키워드 수', f"{search_perf['unique_keywords']}개"],
        ['', ''],
        ['💡 주요 인사이트', '']
    ]
    
    # 인사이트 추가
    for insight in insights:
        summary_data.append([insight, ''])
    
    summary_df = pd.DataFrame(summary_data, columns=['항목', '값'])
    summary_df.to_excel(writer, sheet_name='📊 주간 성과 요약', index=False)


def apply_excel_formatting(filename):
    """Excel 파일 스타일링 적용"""
    
    wb = openpyxl.load_workbook(filename)
    
    # 색상 정의
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    summary_fill = PatternFill(start_color='E7F3FF', end_color='E7F3FF', fill_type='solid')
    
    # 폰트 정의
    header_font = Font(color='FFFFFF', bold=True, size=12)
    title_font = Font(bold=True, size=14)
    
    # 테두리 정의
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # 헤더 스타일링
        if ws.max_row > 0:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
        
        # 요약 시트 특별 처리
        if '요약' in sheet_name:
            for row_num in range(1, ws.max_row + 1):
                cell_a = ws[f'A{row_num}']
                cell_b = ws[f'B{row_num}']
                
                # 섹션 제목 스타일링
                if cell_a.value and ('📊' in str(cell_a.value) or '🔍' in str(cell_a.value) or '💡' in str(cell_a.value)):
                    cell_a.font = title_font
                    cell_a.fill = summary_fill
                
                # 테두리 적용
                cell_a.border = thin_border
                cell_b.border = thin_border
        
        # 열 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # 차트 추가 (상위 3개 시트에만)
    add_charts_to_workbook(wb)
    
    wb.save(filename)


def add_charts_to_workbook(wb):
    """워크북에 차트 추가"""
    
    try:
        # UTM 캠페인 성과 시트에 차트 추가
        if 'UTM 캠페인 성과' in wb.sheetnames:
            ws = wb['UTM 캠페인 성과']
            
            if ws.max_row > 1:
                # 세션 기준 상위 10개 캠페인 바차트
                chart = BarChart()
                chart.title = "상위 10개 캠페인 세션 수"
                chart.x_axis.title = "캠페인"
                chart.y_axis.title = "세션"
                
                # 데이터 범위 (상위 10개만)
                max_rows = min(ws.max_row, 11)
                data = Reference(ws, min_col=5, min_row=1, max_row=max_rows)  # Sessions 열
                cats = Reference(ws, min_col=1, min_row=2, max_row=max_rows)  # Campaign 열
                
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                
                ws.add_chart(chart, "I2")
        
        # 채널별 성과 시트에 파이차트 추가
        if '채널별 성과' in wb.sheetnames:
            ws = wb['채널별 성과']
            
            if ws.max_row > 1:
                chart = PieChart()
                chart.title = "채널별 세션 점유율"
                
                data = Reference(ws, min_col=3, min_row=1, max_row=ws.max_row)  # Sessions 열
                cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)  # Channel 열
                
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                
                ws.add_chart(chart, "G2")
    
    except Exception as e:
        print(f"차트 생성 중 오류 발생: {e}")


def create_quick_summary_report(utm_data, gsc_data):
    """간단한 요약 보고서 생성 (테스트용)"""
    
    filename = f"quick_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        
        # UTM 데이터 요약
        utm_summary = utm_data.groupby(['Campaign', 'Source', 'Medium']).agg({
            'Sessions': 'sum',
            'Users': 'sum',
            'Conversions': 'sum'
        }).sort_values('Sessions', ascending=False).reset_index()
        
        utm_summary.to_excel(writer, sheet_name='UTM 요약', index=False)
        
        # 검색 데이터
        gsc_data.head(20).to_excel(writer, sheet_name='검색 키워드', index=False)
    
    return filename


if __name__ == "__main__":
    print("보고서 생성 모듈 테스트...")
    
    # 테스트용 더미 데이터
    test_utm = pd.DataFrame({
        'Campaign': ['summer_sale', 'brand_awareness'],
        'Source': ['google', 'facebook'], 
        'Medium': ['cpc', 'social'],
        'Sessions': [1200, 600],
        'Users': [1000, 500],
        'Conversions': [50, 15]
    })
    
    test_gsc = pd.DataFrame({
        'Query': ['best shoes', 'running shoes'],
        'Clicks': [100, 80],
        'Impressions': [2000, 1500],
        'CTR': [0.05, 0.053],
        'Position': [3.2, 4.1]
    })
    
    filename = create_quick_summary_report(test_utm, test_gsc)
    print(f"테스트 보고서 생성 완료: {filename}")